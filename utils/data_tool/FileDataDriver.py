#!/usr/bin/env python3 
# -*- coding: utf-8 -*-
"""
@Date    : 2025/5/10 21:29
@Author  : Poco Ray
@File    : FileDataDriver.py
@Software: PyCharm
@Desc    : Description
"""
import openpyxl
import yaml
from config.global_config import EXCEL_PATH, SHEET_NAME, YAML_PATH
from utils.encrypt_tool.AES import ENCRYPT_AES
from utils.encrypt_tool.RSA import ENCRYPT_RSA


class FileReader:

    @staticmethod
    def read_excel_to_dict(file_path: str = EXCEL_PATH, sheet_name: str = SHEET_NAME) -> list[dict[str, str]]:
        """
        读取Excel文件并转换为字典列表
        :param file_path: Excel文件路径
        :param sheet_name: 工作表名称
        :return: 转换后的字典列表，[{}, {}, ...]
        """
        # 1. 加载Excel文件
        try:
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        # 2. 获取工作表
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.create_sheet(sheet_name)

        # 3. 读取数据
        data_for_list = []
        headers = [cell.value for cell in worksheet[2]]
        for row in worksheet.iter_rows(min_row=3, values_only=True):
            data_for_list.append(dict(zip(headers, row)))
        workbook.close()
        return data_for_list

    @staticmethod
    def write_excel_from_data(
            file_path: str = EXCEL_PATH, sheet_name: str = SHEET_NAME,
            row=None, column=None, value=None):
        """
        将数据写入Excel文件
        :param file_path: Excel文件路径
        :param sheet_name: 工作表名称
        :param row: 行号
        :param column: 列号
        :param value: 写入的值
        """
        # 1. 加载Excel文件
        try:
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        # 2. 获取工作表
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.create_sheet(sheet_name)

        # 3. 写入数据
        worksheet.cell(row=row, column=column, value=value)
        workbook.save(file_path)
        workbook.close()

    @staticmethod
    def read_yaml_to_dict(file_path: str = YAML_PATH) -> list[dict]:
        """
        读取YAML文件并转换为字典列表
        :param file_path: YAML文件路径
        :return: 转换后的字典列表
        """
        with open(file_path, 'r', encoding='utf-8') as file:
            data_for_list = yaml.safe_load(file)
        return data_for_list

    @staticmethod
    def write_data_to_yaml(data=None, file_path: str = YAML_PATH):
        """
        将数据写入YAML文件
        :param file_path: YAML文件路径
        :param data: 要写入的数据
        """
        with open(file_path, 'w', encoding='utf-8') as file:
            yaml.dump(data, file, allow_unicode=True)

    @staticmethod
    def replace_variable_in_yaml(data, var):
        """
        替换YAML文件中的变量, 把data参数中的 {{var}} 转换为对应的变量值
        :param data: YAML数据
        :param var: 要替换的变量
        :return: 替换后的数据
        """
        if isinstance(data, str):
            # 替换字符串中的变量
            for key, value in var.items():
                value = str(value)  # 避免int类型数据导致replace方法失效
                data = data.replace('{{' + key + '}}', value)
            return data
        elif isinstance(data, list):
            return [FileReader.replace_variable_in_yaml(item, var) for item in data]
        elif isinstance(data, dict):
            return {key: FileReader.replace_variable_in_yaml(value, var) for key, value in data.items()}
        else:
            return data

    @staticmethod
    def data_encrypt_by_aes(data: dict):
        """
        将传过来的数据进行AES加密并返回, 需要加密的字段以@标识, 若没有@标识则返回原数据
        :param data: 要加密的数据
        :return: 加密后的数据
        :Usage:
            data = {"@username": "pocoray", "@password": "123456"}
            result = FileReader.data_encrypt_by_aes(data)
            print(result)  # {'username': 'k4FQF5PvWs8d+AZKnu1w6g==', 'password': 'mdSm0RmB+xAKrTah3DG31A=='}
        """
        new_data = {}
        # 1. 遍历确定key的第一个字符是否为@, 是则加密
        # 2. 传过来的数据可能非dict类型
        try:
            for key in data:
                if key[0] == '@':
                    new_data[key[1:]] = ENCRYPT_AES.encrypt(data[key])
                else:
                    new_data[key] = data[key]
            return new_data
        except:
            return None


if __name__ == '__main__':
    data1 = FileReader.read_excel_to_dict()
    print(data1)
    data2 = FileReader.read_yaml_to_dict()
    print(data2)
    all_extract = {"token": "1234567890"}
    Params = {'application': 'app', 'application_client_type': 'weixin', 'token': '{{token}}'}
    result = FileReader.replace_variable_in_yaml(Params, all_extract)
    print(result)
    data3 = {"@username": "pocoray", "@password": "123456"}
    result = FileReader.data_encrypt_by_aes(data3)
    print(result)
